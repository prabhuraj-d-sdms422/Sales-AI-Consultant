"""Lead tracker.

V1 default: append to local Excel (data/leads.xlsx).
Optional: append the same row to Google Sheets when configured.
"""

import asyncio
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config.settings import settings
from app.models.state import ConversationState

EXCEL_PATH = "data/leads.xlsx"
OLD_HEADERS = [
    "Timestamp",
    "Lead Temp",
    "Name",
    "Company",
    "Email",
    "Phone",
    "Industry",
    "Problem",
    "Budget Signal",
    "Urgency",
    "Decision Maker",
    "Solutions Discussed",
    "Objections Raised",
    "Stage",
    "Session ID",
]
HEADERS = [
    "Timestamp",
    "Lead Temp",
    "Name",
    "Company",
    "Email",
    "Phone",
    "Industry",
    "Problem",
    "Budget Signal",
    "Urgency",
    "Decision Maker",
    "Solutions Discussed",
    "Objections Raised",
    "Stage",
    "Session ID",
    "HubSpot URL",
]


def _build_row(state: ConversationState) -> list[str]:
    profile = state.get("client_profile", {}) or {}
    return [
        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        str(state.get("lead_temperature", "cold")).upper(),
        str(profile.get("name", "") or ""),
        str(profile.get("company", "") or ""),
        str(profile.get("email", "") or ""),
        str(profile.get("phone", "") or ""),
        str(profile.get("industry", "") or ""),
        str(profile.get("problem_understood") or profile.get("problem_raw", "") or ""),
        str(profile.get("budget_signal", "") or ""),
        str(profile.get("urgency", "") or ""),
        str(profile.get("decision_maker", "")),
        ", ".join(state.get("solutions_discussed", []) or []),
        ", ".join(state.get("objections_raised", []) or []),
        str(state.get("conversation_stage", "") or ""),
        str(state.get("session_id", "") or ""),
        str(state.get("hubspot_contact_url", "") or ""),
    ]


async def append_lead_locally(state: ConversationState) -> None:
    """Append a lead row to local Excel (always-on)."""
    os.makedirs("data", exist_ok=True)
    row = _build_row(state)
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        # Auto-upgrade headers for existing workbooks created before HubSpot URL column existed.
        try:
            existing = [c.value for c in ws[1] if c.value is not None]
            if existing == OLD_HEADERS:
                for col_idx, header in enumerate(HEADERS, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
        except Exception:
            pass
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
    ws.append(row)
    wb.save(EXCEL_PATH)


def _append_google_sheet_sync(row: list[str]) -> None:
    """Sync append into Google Sheets. Intended to run in a thread."""
    from google.oauth2 import service_account  # noqa: PLC0415
    from googleapiclient.discovery import build  # noqa: PLC0415

    if not settings.google_sheets_spreadsheet_id.strip():
        return

    creds_path = settings.google_sheets_credentials_file.strip()
    if not creds_path:
        return
    # Resolve relative paths from repo root so running uvicorn from backend/ works.
    p = Path(creds_path)
    if not p.is_absolute():
        p = Path(settings.repo_root) / creds_path

    creds = service_account.Credentials.from_service_account_file(
        str(p),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)

    sheet = settings.google_sheets_worksheet_name or "Leads"
    range_a1 = f"{sheet}!A:Z"

    # Prevent Google Sheets from trying to parse values (e.g. "+91..." as a formula).
    safe_row = list(row)
    if len(safe_row) > 5 and isinstance(safe_row[5], str) and safe_row[5].lstrip().startswith("+"):
        safe_row[5] = "'" + safe_row[5].lstrip()

    # Ensure headers exist if the sheet is empty.
    header_range = f"{sheet}!A1:Z1"
    existing = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=settings.google_sheets_spreadsheet_id, range=header_range)
        .execute()
        .get("values", [])
    )
    if not existing:
        service.spreadsheets().values().update(
            spreadsheetId=settings.google_sheets_spreadsheet_id,
            range=f"{sheet}!A1",
            valueInputOption="RAW",
            body={"values": [HEADERS]},
        ).execute()
    else:
        # Auto-upgrade existing trackers created before HubSpot URL column existed.
        # Only overwrite row 1 when it exactly matches the previous header set.
        try:
            row1 = [str(x) for x in (existing[0] or [])]
            if row1 == OLD_HEADERS:
                service.spreadsheets().values().update(
                    spreadsheetId=settings.google_sheets_spreadsheet_id,
                    range=f"{sheet}!A1",
                    valueInputOption="RAW",
                    body={"values": [HEADERS]},
                ).execute()
        except Exception:
            pass

    service.spreadsheets().values().append(
        spreadsheetId=settings.google_sheets_spreadsheet_id,
        range=range_a1,
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": [safe_row]},
    ).execute()


async def append_lead_google_sheets(state: ConversationState) -> None:
    """Append a lead row to Google Sheets when configured; no-op otherwise."""
    if not settings.google_sheets_enabled:
        return
    if not settings.google_sheets_spreadsheet_id.strip():
        return
    if not settings.google_sheets_credentials_file.strip():
        return

    row = _build_row(state)
    await asyncio.to_thread(_append_google_sheet_sync, row)
